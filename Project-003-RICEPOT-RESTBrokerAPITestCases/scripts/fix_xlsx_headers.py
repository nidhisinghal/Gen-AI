from openpyxl import load_workbook
from pathlib import Path

p = Path(r"c:\Users\Anuj\Downloads\AI Tester\Project\Project-003-RICEPOT-RESTBrokerAPITestCases\Rest Broker API Test Case Sheet.xlsx")
if not p.exists():
    raise SystemExit(f"File not found: {p}
")

wb = load_workbook(p)
ws = wb.active

headers = [
    "Scenario TID",
    "TestCase Description",
    "PreCondition",
    "TestSteps",
    "Expected Result",
    "Actual Result",
    "Steps to Execute",
    "Expected Result",
    "Actual Result",
    "Status",
    "Executed QA Name",
    "Misc (Comments)",
    "Priority",
    "Is Automated",
]

# Write headers exactly into first row
for col_index, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_index, value=header)

wb.save(p)
print(f"Fixed headers in: {p}")
